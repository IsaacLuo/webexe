from openpyxl import load_workbook,Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import sys
def write_plate_definition(filename, plate):
    wb = Workbook()
    main_sheet = wb['Sheet']
    for i in range(1,plate['cols']+1):
        main_sheet[chr(65+i)+'1'] = i
    for i in range(1,plate['rows']+1):
        main_sheet['A{}'.format(i+1)] = chr(ord('A')+i-1)
    cols = plate['cols']
    for i in range(plate['type']):
        sheet_row = i//cols + 2
        sheet_col = chr(i%cols + 1 + ord('A'))
        if plate['content'][i]:
            main_sheet['{}{}'.format(sheet_col,sheet_row)].value = plate['content'][i]['text']
            if 'fill' in plate['content'][i]:
                main_sheet['{}{}'.format(sheet_col,sheet_row)].fill = PatternFill('solid', fgColor=plate['content'][i]['fill'])

    for col in main_sheet.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        main_sheet.column_dimensions[column].width = adjusted_width

    wb.save(filename=filename)