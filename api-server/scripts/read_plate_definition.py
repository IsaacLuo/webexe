# coding: utf-8
import re
import sys
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Color, PatternFill, Font, Border

# A1 -> 0
def well_name_to_id(well_name, cols=12):
    try:
        match = re.match(r'([A-Za-z]+)(\d+)', well_name.upper())
        row = ord(match[1]) - ord('A')
        return row*cols+int(match[2])-1
    except:
        print('error in analysing {}'.format(well_name), file=sys.stderr, flush=True)
        return 0

# 0 -> A1
def id_to_well_name(id, cols=12):
    col = id % cols
    row = id // cols
    return '{}{}'.format(chr(ord('A')+row),col+1)

def read_plate_definition(filename):
    wb = load_workbook(filename)
    sheet_names = wb.sheetnames
    sheet = wb[sheet_names[0]]
    rows = sheet.max_row - 1
    cols = sheet.max_column - 1
    plate_type = rows*cols
    if plate_type != 96 and plate_type != 384:
        raise RuntimeError('plate type neither 96 nor 384, it has {} rows and {} cols'.format(rows, cols))

    content = []
    for i in range(rows):
        for j in range(cols):
            cell = sheet['{}{}'.format(chr(ord('B')+j), i+2)]
            if cell:
                content.append({'text':cell.value})
            else:
                content.append(None)
            
    plate = {
        'type': plate_type,
        'rows': rows,
        'cols': cols,
        'content': content,
    }
    return plate